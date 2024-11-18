# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import logging
import re
from sqlite3 import IntegrityError

import scrapy
# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from dateparser import parse
from datetime import datetime
import sqlite3
import pandas as pd
from openpyxl.reader.excel import load_workbook
from scrapy.crawler import Crawler

from scrapy.exceptions import DropItem
from unicodedata import category


class JobPipeline:

    def process_item(self, item, spider):
        item['url'] = item['url'].strip() if item.get('url') else None
        item['title'] = item['title'].strip() if item.get('title') else None
        item['company'] = item['company'].strip() if item.get('company') else None
        item['published_date'] = self.convert_date(item['published_date']) if item.get('published_date') else None
        item['apply_date'] = self.convert_date(item['apply_date']) if item.get('apply_date') else None
        item['location'] = item['location'].strip() if item.get('location') else None
        item['category'] = ', '.join([c.strip() for c in item['category']]) if item.get('category') else None
        # if job_type from HTML
        # item['job_type'] = ', '.join([t.strip() for t in item['job_type']]) if item.get('job_type') else None
        # if job_type from JSON
        item['job_type'] = item['job_type'].strip() if item.get('job_type') else None
        # if description from JSON with HTML tags
        # item['description'] = item['description'].strip() if item.get('description') else None
        # if description from HTML and not from JSON
        item['description'] = '\n'.join([d.strip() for d in item['description']]) if item.get('description') else None
        item['processed_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        item['phone'] = item['phone'].strip().replace(" ", "-") if item.get('phone') else None
        item['email'] = item['email'].strip() if item.get('email') else None

        if item["description"]:
            item['additional_contacts'] = self.extract_contacts(item["description"])

        return item

    @staticmethod
    def convert_date(swedish_date: str) -> str | None:
        try:
            if len(swedish_date.split()) == 2 and ":" not in swedish_date:
                swedish_date = f"{swedish_date} {datetime.now().year}"
            date_obj = parse(swedish_date, languages=['sv'])
            return date_obj.strftime("%Y-%m-%d")
        except (ValueError, AttributeError):
            return None

    @staticmethod
    def extract_contacts(text: str) -> str | None:
        text = text.lower()
        email_pattern = r"[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*@(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?"
        emails = re.findall(email_pattern, text)
        phone_pattern = r"\+?\d{1,4}[-.\s]?\(?\d{1,3}\)?[-.\s]?(?:\d[-.\s]?){7,13}\d"
        phones = re.findall(phone_pattern, text)
        output = []
        if emails:
            output.append(f"Email: {', '.join(emails)}")
        if phones:
            output.append(f"Phones: {', '.join(phones)}")

        return "\n".join(output) if output else None


class DatabasePipeline:
    def __init__(self, crawler):
        self.connection = crawler.db_connection
        self.settings = crawler.settings
        self.item_counter = 0

        self.batch_size = 50
        self.cursor = self.connection.cursor()

    @classmethod
    def from_crawler(cls, crawler):
        return cls(crawler)

    def process_item(self, item, spider):
        try:
            self.cursor.execute(
                '''
                INSERT INTO jobs (
                url, title, company, published_date, apply_date, location, category, job_type, description, 
                processed_date, phone, email, additional_contacts
                ) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    item.get('url'), item.get('title'), item.get('company'),
                    item.get('published_date'), item.get('apply_date'), item.get('location'),
                    item.get('category'), item.get('job_type'), item.get('description'), item.get('processed_date'),
                    item.get('phone'), item.get('email'), item.get('additional_contacts')
                )
            )
            self.connection.commit()

        except IntegrityError:
            spider.logger.info(f"Drop item {item['url']}. URL already exists in the database.")
            raise DropItem()
        else:
            self.item_counter += 1
            if self.item_counter % self.batch_size == 0:
                spider.logger.info(f"~~~Added {self.item_counter} jobs")
        return item

    def close_spider(self, spider):
        self.cursor.close()


class ExcelSavePipeline:
    """Save the bunch of items to Excel file"""

    def __init__(self):
        self.items = []
        self.batch_size = 50
        self.excel_file = None
        self.sheet_name = 'Sheet1'
        self.start_row = 0
        self.workbook = None


    def open_spider(self, spider: scrapy.Spider):
        self.excel_file = spider.settings.get("EXCEL_FILE_INCREMENTAL")
        self._init_workbook()

    def _init_workbook(self):
        """The function gets the number of rows in an Excel file or creates it if it does not exist."""
        try:
            self.workbook = load_workbook(self.excel_file)
            if self.sheet_name in self.workbook.sheetnames:
                self.start_row = self.workbook[self.sheet_name].max_row
        except FileNotFoundError:
            self.workbook = None
            self.start_row = 0

    def process_item(self, item, spider):
        self.items.append(item)
        if len(self.items) >= self.batch_size:
            self.save_data_to_excel()
        return item

    def save_data_to_excel(self):

        df = pd.DataFrame(self.items)
        header_enabled = False if self.start_row else True

        writer_params = {
            'engine': 'openpyxl',
            'mode': 'a' if self.start_row else 'w'
        }
        if self.start_row > 0:
            writer_params['if_sheet_exists'] = "overlay"

        with pd.ExcelWriter(self.excel_file, **writer_params) as writer:
            df.to_excel(writer, sheet_name=self.sheet_name, index=False, header=header_enabled, startrow=self.start_row)
            self.start_row += len(df)

        self.items.clear()

    def close_spider(self, spider):
        if self.items:
            self.save_data_to_excel()


class ExcelFinalExportPipeline:
    """
    Save the all data to xlsx
    """

    def __init__(self):
        self.excel_file = None

    def open_spider(self, spider: scrapy.Spider):
        self.excel_file = spider.settings.get("EXCEL_FILE_FROM_DB")

    def close_spider(self, spider):
        spider.logger.info(f"Start saving all records to {self.excel_file}")
        connection = spider.crawler.db_connection
        # Создание индекса, если нужно
        cursor = connection.cursor()
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_company ON jobs(company);')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_published_date ON jobs(published_date DESC);')
        connection.commit()
        query = '''
        SELECT j.*, c.company_jobs_in_db 
        FROM jobs j
        LEFT JOIN (
            SELECT company, COUNT(*) AS company_jobs_in_db
            FROM jobs
            GROUP BY company
        ) c ON j.company = c.company
        ORDER BY j.published_date DESC;
        '''
        df = pd.read_sql_query(query, connection)
        cursor.execute('DROP INDEX IF EXISTS idx_company;')
        cursor.execute('DROP INDEX IF EXISTS idx_published_date;')
        connection.commit()
        record_count = df.shape[0]
        spider.logger.info(f"Total record count in DB {record_count}")
        df.to_excel(self.excel_file, index=False)
        spider.logger.info(f"All records are saved to {self.excel_file}")
